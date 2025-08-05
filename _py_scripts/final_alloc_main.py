# FINAL ALLOCATIONS - Dynamic Template
import os
import pandas as pd
from openpyxl import Workbook
import glob

# Path Mgmt & core defs/lists <<<
#▽―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---―――――――▽
#region
script_dir = os.path.dirname(__file__)
parent_dir = os.path.abspath(os.path.join(script_dir, os.pardir))
x_template_path = os.path.join(parent_dir, '_templates', 'alloc_template.csv')
output_tgt = os.path.join(parent_dir, 'FINAL_Allocations.xlsx')

print('')
print('=== Begin Allocations Process ===')
print(f"Template path: {x_template_path}")
print(f"Output target: {output_tgt}")
print('')

file_src = os.path.join(parent_dir, 'Timesheet_Set')
file_src_templates = os.path.join(parent_dir, '_templates')
file_src_payroll = os.path.join(parent_dir, 'READ_Payroll_Report')

try:
    df_xRs1 = pd.read_csv(x_template_path, header=0)
    print("Template loaded successfully")
    print('.'*28)
except (FileNotFoundError, ValueError, PermissionError, OSError) as e:
    print(f"Error loading template: {e}")
    df_xRs1 = None

def fill_constants(df: pd.DataFrame) -> None:
    constants = {
        'CurrencyCode': 'USD',
        'ExchRate': '1',
        'MainAccount': '153115',
        'DIVISION': 'A',
        'ENTITY': '21',
        'COUNTRY': 'US',
        'INTERCOMPANY': '99',
    }

    for col, value in constants.items():
        if col in df.columns:
            df[col] = value

sheet_names: list[str] = [
    "Voucher_xx-xx-xx",
    "Approval Support",
    "Editorial - 291",
    "Product Devel - 294",
    "Photo - 295",
    "Prepress - 297",
    "Design - 304",
    "CAM-WH - 305",
    "DIGITAL HOURS REALLOCATE",
    "Payroll Allocation Report",
    "Payroll Alloc_PIVOT",
]

file_mapping: dict[str, str] = {
    os.path.join(file_src, "Blockchain_8Codes.csv"): "DIGITAL HOURS REALLOCATE",
}

file_mapping_depts: dict[str, str] = {
    os.path.join(file_src, "Editorial Garth Group.csv"): "Editorial - 291",
    os.path.join(file_src, "ProdDev_Merged.csv"): "Product Devel - 294",
    os.path.join(file_src, "Photo Timesheets.csv"): "Photo - 295",
    os.path.join(file_src, "Pre-Press.csv"): "Prepress - 297",
    os.path.join(file_src, "Editorial Luis Group.csv"): "CAM-WH - 305",
}

#endregion
#△――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---――――――――――――――△

# Merge ProdDev & Ent. <<<
#▽―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---―――――――▽
#region

def get_merged_proddev_data():
    """
    Generator that initializes merged Product Development + Entertainment data
    without overwriting source files.
    """
    try:
        print("Generating merged Product Development + Entertainment data...")

        # load sources
        proddev_path = os.path.join(file_src, "ProdDev_Merged.csv")
        ent_path = os.path.join(file_src, "Entertainment.csv")

        if not os.path.exists(proddev_path):
            print(f"Warning: {proddev_path} not found")
            return None

        if not os.path.exists(ent_path):
            print(f"Warning: {ent_path} not found")
            return None

        df_proddev = pd.read_csv(proddev_path)
        df_ent = pd.read_csv(ent_path)

        # Prepare DataFrames by renaming Total columns
        df_proddev = df_proddev.rename(columns={'Total': 'Total_ProdDev'})
        df_ent = df_ent.rename(columns={'Total': 'Total_ENT'})

        # Merge df's on common cols (including Sport to avoid suffixes)
        merge_cols = ['Collection Code', 'AX Project Code', 'Collection Description', 'Sport']
        # Only use columns that exist in both dataframes
        available_merge_cols = [col for col in merge_cols if col in df_proddev.columns and col in df_ent.columns]

        if not available_merge_cols:
            print("Warning: No common merge columns found. Concatenating instead...")
            # If no common cols, just concatenate
            merged_df = pd.concat([df_proddev, df_ent], ignore_index=True)
            merged_df['Total'] = merged_df[['Total_ProdDev', 'Total_ENT']].fillna(0).sum(axis=1)
        else:
            # Merge on available cols (Sport included prevents suffix issues)
            merged_df = df_proddev.merge(df_ent, on=available_merge_cols, how='outer')

            # Fill NaN values in Total cols with 0 for sum
            total_cols = ['Total_ProdDev', 'Total_ENT']
            merged_df[total_cols] = merged_df[total_cols].fillna(0)
            merged_df['Total'] = merged_df[total_cols].sum(axis=1)

        # Rename for clarity
        merged_df = merged_df.rename(columns={
            'Total_ProdDev': '_ProdDev_Original',
            'Total_ENT': '_Entertainment_Merged'
        })

        # Put Total in the right position
        if 'Total' in merged_df.columns:
            total_col = merged_df.pop('Total')
            merged_df.insert(len(merged_df.columns), 'Total', total_col)

        # Handle Total row repositioning (similar to editorial merge)
        total_row_index = merged_df[(merged_df.iloc[:, 0] == 'Total') |
                                   (merged_df.iloc[:, 1] == 'Total')].index
        if not total_row_index.empty:
            total_row = merged_df.loc[total_row_index].copy()
            merged_df = merged_df.drop(total_row_index)
            new_index = min(total_row_index[0] + 3, len(merged_df))
            upper_half = merged_df.iloc[:new_index]
            lower_half = merged_df.iloc[new_index:]
            merged_df = pd.concat([upper_half, total_row, lower_half]).reset_index(drop=True)

        # Drop rows where all values are NaN
        merged_df = merged_df.dropna(how='all')

        # Limit to reasonable number of rows
        merged_df = merged_df.iloc[:350]

        print("Successfully generated merged Product Development + Entertainment data")
        return merged_df

    except Exception as e:
        print(f"Error in get_merged_proddev_data: {e}")
        return None

#endregion
#△――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---――――――――――――――△

def convert_to_number(value):
    try:
        return float(value)
    except ValueError:
        return value

# Create Pivot Table <<<
#▽―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---―――――――▽
#region
def emulate_pivot():
    """
    **Emulate the legacy excel pivot table for 'Payroll Allocation Report.csv' file**
    In the legacy pivot table:
    1.) Rows were ['Cost Center']
    2.) Columns were ['Division2']
    3.) Values were Sum of ['Amount in transaction currency']
    **Utilizing Python**
    4.) Normalize data by cleaning text fields and handling Excel error values
    5.) Create pivot table with grand totals for both rows and columns
    6.) Return pivot table dataframe for direct use in Excel workbook
    7.) Show progress output in terminal throughout processing
    """
    # Use glob to find any XLSX file in the payroll directory
    payroll_dir = os.path.join(parent_dir, 'READ_Payroll_Report')
    xlsx_files = glob.glob(os.path.join(payroll_dir, "*.xlsx"))

    if not xlsx_files:
        print(f"Error: No XLSX files found in {payroll_dir}")
        print("Please place a payroll report XLSX file in the READ_Payroll_Report directory")
        return None

    # Use the first XLSX file found
    src_path = xlsx_files[0]
    xlsx_filename = os.path.basename(src_path)

    try:
        print(f"Loading Payroll Allocation Report: {xlsx_filename}...")

        # Read the XLSX file
        df = pd.read_excel(src_path)
        print(f"Loaded {len(df)} rows of data")        # Data normalization
        print("Normalizing data...")
        for col in df.columns:
            if df[col].dtype == 'object':  # Only process text/mixed columns
                df[col] = df[col].astype(str).str.replace(',', '').str.strip() # Convert to string, remove commas, strip whitespace
                # Replace common Excel error values and empty strings with appropriate defaults
                df[col] = df[col].replace(['', '-', '- ', 'NaN', 'nan', '#DIV/0!', '#N/A', '#VALUE!', '$- ', 'None'], '')
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                if not numeric_col.isna().all():  # If any values converted successfully
                    df[col] = numeric_col.fillna(0)  # Fill NaN with 0 for numeric columns

        # Create pivot table
        print("Creating pivot table...")
        pivot_table = df.pivot_table(
            index='Cost Center',
            columns='Division2',
            values='Amount in transaction currency',
            aggfunc='sum',
            fill_value=0
        )

        print(f"Pivot table created with {len(pivot_table)} cost centers and {len(pivot_table.columns)} divisions")

        # Add Grand Total row and column
        pivot_table.loc['Grand Total'] = pivot_table.sum()
        pivot_table['Grand Total'] = pivot_table.sum(axis=1)  # Add column-wise totals

        print("Pivot table processing complete!")
        print('.' * 32, '\n')
        return pivot_table

    except FileNotFoundError:
        print(f"Error: Could not find source file at {src_path}")
        return None
    except KeyError as e:
        print(f"Error: Missing expected column {e} in the data")
        return None
    except Exception as e:
        print(f"Error processing pivot table: {e}")
        return None

if df_xRs1 is not None:
    # Generate pivot table for Payroll Alloc_PIVOT worksheet
    print("\n=== Generating Pivot Table ===")
    pivot_df = emulate_pivot()
    if pivot_df is None:
        print("Warning: Pivot table generation failed. The Payroll Alloc_PIVOT sheet may be empty.")

#endregion
#△――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---――――――――――――――△

# >>> Setting up the workbook and core column set <<<
#▽―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---―――――――▽
#region
    # Create a single empty row if template is empty
    if len(df_xRs1) == 0:
        df_xRs1 = pd.DataFrame([[''] * len(df_xRs1.columns)], columns=df_xRs1.columns)
    df_output = pd.concat([df_xRs1] * 400, ignore_index=True)

    fill_constants(df_output)

    # Convert all values to numbers where possible
    df_output = df_output.map(convert_to_number)

    wb = Workbook()
    excel_path = output_tgt
    for sheet_name in sheet_names: # Create worksheets and populate data
        ws = wb.create_sheet(title=sheet_name)

        # Check if this sheet should be populated from a specific CSV file
        csv_file_path = None
        for file_path, mapped_sheet in file_mapping.items():
            if mapped_sheet == sheet_name:
                csv_file_path = file_path
                break

        # Check if this sheet should be populated from department CSV files
        dept_csv_file_path = None
        for file_path, mapped_sheet in file_mapping_depts.items():
            if mapped_sheet == sheet_name:
                dept_csv_file_path = file_path
                break

        if csv_file_path and os.path.exists(csv_file_path):
            # Populate from specific CSV file
            try:

                # Special handling for DIGITAL HOURS REALLOCATE - has 2 header rows
                if sheet_name == "DIGITAL HOURS REALLOCATE":
                    # Read the CSV without skipping any rows to get both header rows
                    df_with_headers = pd.read_csv(csv_file_path, header=None)

                    # Extract the first row (department names) and second row (column names)
                    dept_headers = df_with_headers.iloc[0].tolist()
                    col_headers = df_with_headers.iloc[1].tolist()

                    # Read the data starting from row 3 onwards, using row 2 as column names
                    df_specific = pd.read_csv(csv_file_path, header=1)  # Use row 2 as header

                    # Drop any rows that are actually the header rows (if they got included)
                    if len(df_specific) > 0:
                        # Remove any rows that might be duplicate headers
                        first_col_name = df_specific.columns[0] if len(df_specific.columns) > 0 else ""
                        df_specific = df_specific[df_specific.iloc[:, 0] != first_col_name]

                    # Write both header rows to Excel
                    # First row: department names
                    for col_idx, dept_name in enumerate(dept_headers, start=1):
                        ws.cell(row=1, column=col_idx, value=dept_name)

                    # Second row: column names
                    for col_idx, col_name in enumerate(col_headers, start=1):
                        ws.cell(row=2, column=col_idx, value=col_name)

                    # Find the rightmost 'Total' column (should be 'Total.9' which is the sum column)
                    total_columns = [col for col in df_specific.columns if col.startswith('Total')]
                    if total_columns:
                        # Use the last 'Total' column (Total.9 - the Entertainment total)
                        total_col_name = total_columns[-1]  # This should be 'Total.9'
                    else:
                        # Fallback to last column if no 'Total' columns found
                        total_col_name = df_specific.columns[-1]

                    # Convert to numeric first, then filter
                    if total_col_name in df_specific.columns:
                        df_specific[total_col_name] = pd.to_numeric(df_specific[total_col_name], errors='coerce')

                        original_row_count = len(df_specific)
                        # Filter: keep rows where Total > 0 (this will also remove NaN values)
                        df_specific = df_specific[df_specific[total_col_name] > 0]
                        filtered_row_count = len(df_specific)

                        print(f"Filtered {original_row_count - filtered_row_count} rows with zero/null totals from {sheet_name}")

                    # Convert data and populate starting from row 3 (after both headers)
                    df_specific = df_specific.map(convert_to_number)

                    for row_idx, row in enumerate(df_specific.values, start=3):  # Start from row 3
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    df_specific = pd.read_csv(csv_file_path, header=0)
                    df_specific = df_specific.map(convert_to_number)

                    for col_idx, col_name in enumerate(df_specific.columns, start=1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row in enumerate(df_specific.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                print(f"■ Populated {sheet_name} with data from source file")
                print('')
            except Exception as e:
                print(f"Error loading {csv_file_path} for {sheet_name}: {e}")
        elif sheet_name == "Product Devel - 294":

            # Special handling for Product Development: use merged data dynamically
            print("\n=== Merging Department Data ===")
            merged_df = get_merged_proddev_data()
            if merged_df is not None:
                try:
                    # Select only the specified columns
                    required_columns = ['AX Project Code', 'Collection Code', 'Collection Description', 'Total']
                    available_columns = [col for col in required_columns if col in merged_df.columns]

                    if available_columns:
                        df_dept_filtered = merged_df[available_columns]

                        # Filter out rows where key columns are empty/null AND Total is 0
                        key_columns = ['AX Project Code', 'Collection Code', 'Collection Description']
                        existing_key_columns = [col for col in key_columns if col in df_dept_filtered.columns]

                        if existing_key_columns and 'Total' in df_dept_filtered.columns:
                            # Create mask for rows to keep
                            mask = pd.Series([True] * len(df_dept_filtered))

                            for col in existing_key_columns:
                                # Check if column values are empty/null
                                col_empty = (
                                    df_dept_filtered[col].isna() |
                                    df_dept_filtered[col].isin(['', 'NaN', 'nan', 'None']) |
                                    (df_dept_filtered[col] == 0)
                                )
                                # If this is the first column, set the mask
                                if col == existing_key_columns[0]:
                                    empty_key_mask = col_empty
                                else:
                                    # AND with previous columns (all must be empty)
                                    empty_key_mask = empty_key_mask & col_empty

                            # Check if Total is 0
                            total_zero = (df_dept_filtered['Total'] == 0) | df_dept_filtered['Total'].isna()

                            # Keep rows where NOT (all key columns empty AND total is 0) AND Total is not 0
                            mask = ~(empty_key_mask & total_zero) & ~total_zero
                            df_dept_filtered = df_dept_filtered[mask]
                        elif 'Total' in df_dept_filtered.columns:
                            # If key columns don't exist, just filter out rows where Total is 0
                            total_zero = (df_dept_filtered['Total'] == 0) | df_dept_filtered['Total'].isna()
                            df_dept_filtered = df_dept_filtered[~total_zero]

                        # Filter out rows where Collection Code begins with '8'
                        if 'Collection Code' in df_dept_filtered.columns:
                            collection_code_mask = ~df_dept_filtered['Collection Code'].astype(str).str.startswith('8', na=False)
                            df_dept_filtered = df_dept_filtered[collection_code_mask]

                        # Filter out the 'Total' row since it includes '8' codes in its calculation
                        if 'Collection Code' in df_dept_filtered.columns:
                            total_row_mask = ~df_dept_filtered['Collection Code'].astype(str).str.contains('Total', case=False, na=False)
                            df_dept_filtered = df_dept_filtered[total_row_mask]

#endregion
#△――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---――――――――――――――△

# Create the Weighted Averages <<<
#▽―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---―――――――▽
#region

                        # calculate weighted average
                        if 'Total' in df_dept_filtered.columns and 'Collection Code' in df_dept_filtered.columns:
                            # Find the total sum from the row where Collection Code contains 'Total'
                            total_mask = df_dept_filtered['Collection Code'].astype(str).str.contains('Total', case=False, na=False)
                            if total_mask.any():
                                total_sum = df_dept_filtered.loc[total_mask, 'Total'].iloc[0]
                            else:
                                # If no 'Total' row found, use sum of all Total values
                                total_sum = df_dept_filtered['Total'].sum()

                            # Calculate weighted average for each row
                            if total_sum != 0:
                                df_dept_filtered['Weighted Avg'] = df_dept_filtered['Total'] / total_sum
                                # Format Weighted Avg to 4 decimal places
                                df_dept_filtered['Weighted Avg'] = df_dept_filtered['Weighted Avg'].round(4)
                            else:
                                df_dept_filtered['Weighted Avg'] = 0.0000

                        # Apply convert_to_number but preserve Weighted Avg formatting
                        for col in df_dept_filtered.columns:
                            if col != 'Weighted Avg':  # Skip Weighted Avg column to preserve formatting
                                df_dept_filtered[col] = df_dept_filtered[col].apply(convert_to_number)

                        for col_idx, col_name in enumerate(df_dept_filtered.columns, start=1):
                            ws.cell(row=1, column=col_idx, value=col_name)

                        for row_idx, row in enumerate(df_dept_filtered.values, start=2):
                            for col_idx, value in enumerate(row, start=1):
                                # Special formatting for Weighted Avg column
                                if df_dept_filtered.columns[col_idx-1] == 'Weighted Avg':
                                    ws.cell(row=row_idx, column=col_idx, value=round(float(value), 4))
                                else:
                                    ws.cell(row=row_idx, column=col_idx, value=value)

                        print(f"■ Populated {sheet_name} with merged Product Development + Entertainment data\n")
                        print('')
                    else:
                        print(f"Warning: No required columns found in merged data for {sheet_name}")
                except Exception as e:
                    print(f"Error processing merged data for {sheet_name}: {e}")
            else:
                print(f"Warning: Could not generate merged data for {sheet_name}")
        elif sheet_name == "Approval Support":
            # Special handling for Approval Support: load template data directly without filtering
            approval_csv_path = os.path.join(file_src_templates, "approval_support_template.csv")
            if os.path.exists(approval_csv_path):
                try:
                    df_approval = pd.read_csv(approval_csv_path, header=0)
                    df_approval = df_approval.map(convert_to_number)

                    for col_idx, col_name in enumerate(df_approval.columns, start=1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row in enumerate(df_approval.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    print(f"■ Populated {sheet_name} with approval support template data")
                    print('')
                except Exception as e:
                    print(f"Error loading approval support template: {e}")
            else:
                print(f"Warning: Approval support template not found at {approval_csv_path}")
        elif dept_csv_file_path and os.path.exists(dept_csv_file_path):
            # Populate from department CSV file with selected columns only
            try:
                print(f"Processing department file: {os.path.basename(dept_csv_file_path)}")
                df_dept = pd.read_csv(dept_csv_file_path, header=0)

                # Select only the specified columns
                required_columns = ['AX Project Code', 'Collection Code', 'Collection Description', 'Total']
                available_columns = [col for col in required_columns if col in df_dept.columns]

                if available_columns:
                    df_dept_filtered = df_dept[available_columns]

                    # Filter out rows where key columns are empty/null AND Total is 0
                    key_columns = ['AX Project Code', 'Collection Code', 'Collection Description']
                    existing_key_columns = [col for col in key_columns if col in df_dept_filtered.columns]

                    if existing_key_columns and 'Total' in df_dept_filtered.columns:
                        # Create mask for rows to keep
                        mask = pd.Series([True] * len(df_dept_filtered))

                        for col in existing_key_columns:
                            # Check if column values are empty/null
                            col_empty = (
                                df_dept_filtered[col].isna() |
                                df_dept_filtered[col].isin(['', 'NaN', 'nan', 'None']) |
                                (df_dept_filtered[col] == 0)
                            )
                            # If this is the first column, set the mask
                            if col == existing_key_columns[0]:
                                empty_key_mask = col_empty
                            else:
                                # AND with previous columns (all must be empty)
                                empty_key_mask = empty_key_mask & col_empty

                        # Check if Total is 0
                        total_zero = (df_dept_filtered['Total'] == 0) | df_dept_filtered['Total'].isna()

                        # Keep rows where NOT (all key columns empty AND total is 0) AND Total is not 0
                        mask = ~(empty_key_mask & total_zero) & ~total_zero
                        df_dept_filtered = df_dept_filtered[mask]
                    elif 'Total' in df_dept_filtered.columns:
                        # If key columns don't exist, just filter out rows where Total is 0
                        total_zero = (df_dept_filtered['Total'] == 0) | df_dept_filtered['Total'].isna()
                        df_dept_filtered = df_dept_filtered[~total_zero]

                    # Filter out rows where Collection Code begins with '8'
                    if 'Collection Code' in df_dept_filtered.columns:
                        collection_code_mask = ~df_dept_filtered['Collection Code'].astype(str).str.startswith('8', na=False)
                        df_dept_filtered = df_dept_filtered[collection_code_mask]

                    # Filter out the 'Total' row since it includes '8' codes in its calculation
                    if 'Collection Code' in df_dept_filtered.columns:
                        total_row_mask = ~df_dept_filtered['Collection Code'].astype(str).str.contains('Total', case=False, na=False)
                        df_dept_filtered = df_dept_filtered[total_row_mask]

                    # Calculate Weighted Average
                    if 'Total' in df_dept_filtered.columns and 'Collection Code' in df_dept_filtered.columns:
                        # Find the total sum from the row where Collection Code contains 'Total'
                        total_mask = df_dept_filtered['Collection Code'].astype(str).str.contains('Total', case=False, na=False)
                        if total_mask.any():
                            total_sum = df_dept_filtered.loc[total_mask, 'Total'].iloc[0]
                        else:
                            # If no 'Total' row found, use sum of all Total values
                            total_sum = df_dept_filtered['Total'].sum()

                        # Calculate weighted average for each row
                        if total_sum != 0:
                            df_dept_filtered['Weighted Avg'] = df_dept_filtered['Total'] / total_sum
                            # Format Weighted Avg to 4 decimal places
                            df_dept_filtered['Weighted Avg'] = df_dept_filtered['Weighted Avg'].round(4)
                        else:
                            df_dept_filtered['Weighted Avg'] = 0.0000

                    # Apply convert_to_number but preserve Weighted Avg formatting
                    for col in df_dept_filtered.columns:
                        if col != 'Weighted Avg':  # Skip Weighted Avg column to preserve formatting
                            df_dept_filtered[col] = df_dept_filtered[col].apply(convert_to_number)

                    for col_idx, col_name in enumerate(df_dept_filtered.columns, start=1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row in enumerate(df_dept_filtered.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            # Special formatting for Weighted Avg column
                            if df_dept_filtered.columns[col_idx-1] == 'Weighted Avg':
                                ws.cell(row=row_idx, column=col_idx, value=round(float(value), 4))
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=value)

                    print(f"■ Populated {sheet_name} with filtered data from source file")
                    print('')
                else:
                    print(f"Warning: No required columns found in {dept_csv_file_path} for {sheet_name}")
            except Exception as e:
                print(f"Error loading {dept_csv_file_path} for {sheet_name}: {e}")

#endregion
#△――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――---――――――――――――――△

        elif sheet_name == "Payroll Allocation Report":

            # Special handling for Payroll Allocation Report: load raw data directly without filtering
            # Use glob to find any XLSX file in the payroll directory
            payroll_dir = os.path.join(parent_dir, 'READ_Payroll_Report')
            xlsx_files = glob.glob(os.path.join(payroll_dir, "*.xlsx"))

            if xlsx_files:
                payroll_xlsx_path = xlsx_files[0]  # Use the first XLSX file found
                xlsx_filename = os.path.basename(payroll_xlsx_path)

                try:
                    print(f"Loading Payroll Allocation Report: {xlsx_filename}...")
                    df_payroll = pd.read_excel(payroll_xlsx_path, header=0)
                    df_payroll = df_payroll.map(convert_to_number)

                    for col_idx, col_name in enumerate(df_payroll.columns, start=1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row in enumerate(df_payroll.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    print(f"■ Populated {sheet_name} with payroll allocation report data")
                    print('')
                except Exception as e:
                    print(f"Error loading payroll allocation report: {e}")
            else:
                print(f"Warning: No XLSX files found in {payroll_dir}")
                print("Please place a payroll report XLSX file in the READ_Payroll_Report directory")
        elif sheet_name == "Payroll Alloc_PIVOT" and pivot_df is not None:
            # Populate with pivot table data
            try:
                # Reset index to make 'Cost Center' a regular column
                pivot_df_reset = pivot_df.reset_index()
                pivot_df_reset = pivot_df_reset.map(convert_to_number)

                for col_idx, col_name in enumerate(pivot_df_reset.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)

                for row_idx, row in enumerate(pivot_df_reset.values, start=2):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                print(f"■ Populated {sheet_name} with pivot table data")
                print('')
            except Exception as e:
                print(f"Error populating {sheet_name} with pivot data: {e}")
        elif sheet_name == "Voucher_xx-xx-xx":
            # Populate with template data (only for Voucher sheet)
            for col_idx, col_name in enumerate(df_output.columns, start=1):# w df headers
                ws.cell(row=1, column=col_idx, value=col_name)

            for row_idx, row in enumerate(df_output.values, start=2): # w df content
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    if 'Sheet' in wb.sheetnames: # drop default worksheet
        wb.remove(wb['Sheet'])

    def populate_alloc_a_column():
        """
        1.) Adds a new colum named 'Allocation A' to the right-most position for each worksheet in `sheet_names` having the following names:
        "Editorial - 291","Product Devel - 294","Photo - 295","Prepress - 297","Design - 304","CAM-WH - 305"

        """

        cost_center_wksheet_mapping: dict[str,str] = {
            '291': 'Editorial - 291',
            '294': 'Product Devel - 294',
            '295': 'Photo - 295',
            '297': 'Prepress - 297',
            '305': 'CAM-WH - 305',
        }

        # Read the pivot table data from the existing pivot_df
        if pivot_df is not None:
            # Create a dictionary to store cost center amounts
            cost_center_src = {}

            # Extract cost center data from pivot table
            # The pivot table has Cost Center as index (column 0) and amounts in subsequent columns
            for cost_center_code in pivot_df.index:
                cost_center_str = str(cost_center_code)
                if cost_center_str in cost_center_wksheet_mapping:
                    # Get the first amount column (column 1 equivalent - first division column)
                    first_amount_column = pivot_df.columns[0]  # First division column
                    amount = pivot_df.loc[cost_center_code, first_amount_column]
                    cost_center_src[cost_center_str] = amount
        else:
            cost_center_src = {}
            print("Warning: Pivot table not available for cost center lookup")


        for cost_center, worksheet_name in cost_center_wksheet_mapping.items():
            if cost_center in cost_center_src:
                cost_center_amount = cost_center_src[cost_center]

                # Find the worksheet in the workbook
                if worksheet_name in [sheet.title for sheet in wb.worksheets]:
                    ws = wb[worksheet_name]

                    # Find the 'Total' column
                    total_col_idx = None
                    for col_idx in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col_idx).value == 'Total':
                            total_col_idx = col_idx
                            break

                    if total_col_idx:
                        # Add 'Allocation A' header in the next column
                        alloc_col_idx = ws.max_column + 1
                        ws.cell(row=1, column=alloc_col_idx, value='Allocation A')

                        # Calculate and populate 'Allocation A' values for each row
                        for row_idx in range(2, ws.max_row + 1):
                            total_value = ws.cell(row=row_idx, column=total_col_idx).value

                            if total_value is not None and isinstance(total_value, (int, float)) and cost_center_amount != 0:
                                allocation_value = total_value / cost_center_amount
                                # Format to 10 decimal places
                                ws.cell(row=row_idx, column=alloc_col_idx, value=round(allocation_value, 10))
                            else:
                                ws.cell(row=row_idx, column=alloc_col_idx, value=0.0000000000)
                    else:
                        print(f"  ⚠ Warning: 'Total' column not found in {worksheet_name}")
                else:
                    print(f"  ⚠ Warning: Worksheet '{worksheet_name}' not found")
            else:
                print(f"  ⚠ Warning: Cost center {cost_center} amount not found for {worksheet_name}")

        return cost_center_src

    # Call the function to test cost center extraction
    populate_alloc_a_column()

    # Remove Weighted Avg and Allocation A columns from specific worksheets
    # Comment out this block if columns are needed in the future
    target_worksheets = ["Editorial - 291", "Product Devel - 294", "Photo - 295", "Prepress - 297", "Design - 304", "CAM-WH - 305"]
    for worksheet_name in target_worksheets:
        if worksheet_name in [sheet.title for sheet in wb.worksheets]:
            ws = wb[worksheet_name]

            # Find and remove Weighted Avg column
            weighted_avg_col = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == 'Weighted Avg':
                    weighted_avg_col = col_idx
                    break
            if weighted_avg_col:
                ws.delete_cols(weighted_avg_col)

            # Find and remove Allocation A column (check again after potential deletion)
            alloc_a_col = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == 'Allocation A':
                    alloc_a_col = col_idx
                    break
            if alloc_a_col:
                ws.delete_cols(alloc_a_col)

    # Save
    try:
        wb.save(excel_path)
        print('.'*32)
        print(f"Excel file saved successfully to: {excel_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        print(f"Attempted path: {excel_path}")

    print('.'*32)
    print(f"Generated template with {len(df_output)} rows")
    print(f"Created {len(sheet_names)} worksheets")
    print('')
else:
    print("Template could not be loaded. Exiting script.")
