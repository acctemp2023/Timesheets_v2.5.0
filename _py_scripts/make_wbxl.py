import os
import time
import openpyxl # type: ignore
# pylint: disable=W0611 W0702
from openpyxl.utils import get_column_letter # type: ignore # noqa: F401
from openpyxl.styles import numbers # type: ignore # noqa: F401
import pandas as pd

time.sleep(3)

parent_dir = os.path.dirname(os.path.dirname(__file__))

tgt_dir_path = parent_dir
tgt_file_name = 'Excel WB - Dept Hrs.xlsx'
tgt_file_path = os.path.join(tgt_dir_path, tgt_file_name)


wb_files = [
    # 'Blank.csv',
    'Omni_bookmerge_output.csv',
    'Editorial Garth Group.csv',
    'Editorial Luis Group.csv',
    'Photo Timesheets.csv',
    'Pre-Press.csv',
    'ProdDev_Merged.csv',
    'Entertainment.csv',
    'Prod-Dev-Baseball.csv',
    'Prod-Dev-Basketball.csv',
    'Prod-Dev-Football.csv',
    'Prod-Dev-Soccer.csv'

]

sheet_names = {
    # 'Blank': 'Blank',
    'Omni_bookmerge_output': 'Omni Merged',
    'Editorial Garth Group': 'Editorial Garth',
    'Editorial Luis Group': 'Editorial Luis',
    'Photo Timesheets': 'Photo Timesheets',
    'Pre-Press': 'Pre-Press',
    'Entertainment': 'Entertainment',
    'ProdDev_Merged':'PD-Summary',
    'Prod-Dev-Baseball': 'PD-BB-Detail',
    'Prod-Dev-Basketball': 'PD-BK-Detail',
    'Prod-Dev-Football': 'PD-FB-Detail',
    'Prod-Dev-Soccer': 'PD-SC-Detail'
}

# Create a new Excel workbook
wb = openpyxl.Workbook()

# drop default sheet created by excel
wb.remove(wb.active)

# Process each file
for file_name in wb_files:
    # Get the base filename without extension for dict lookup
    base_name = file_name.rsplit('.', 1)[0]

    # Get the corresponding sheet name from dict
    sheet_name = sheet_names.get(base_name, base_name)  # Use base_name as fallback

    # Read the CSV file
    file_path = os.path.join(tgt_dir_path, file_name)
    df = pd.read_csv(file_path)

    # Create a new worksheet
    ws = wb.create_sheet(title=sheet_name)

    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            # Check if the column is "Total_Sum" and convert to numeric if possible
            if df.columns[col_num - 1] == "Total_Sum":
                try:
                    cell_value = float(cell_value)  # make float
                    cell.value = cell_value
                    cell.number_format = '0.00'  # Apply num format
                except ValueError:
                    pass  # Leave as is if conv fails
            # Check if val is numeric & apply num format
            elif isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                cell.number_format = '0.00'

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Save wb
wb.save(tgt_file_path)
print(f"Excel workbook saved as: {tgt_file_path}")
time.sleep(0.25)

if __name__ == '__main__':
    pass